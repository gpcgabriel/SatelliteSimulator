import os
import json
from datetime import datetime
from openpyxl import Workbook

class Controller:    
    def print_buffer(self, buff: str):
        print(f'-----{buff}-----')

    def check_if_path_exists(self, path) -> bool:
        return os.path.exists(path)
    
    def create_path(self, path_name):
        try:
            os.makedirs(path_name)
        except Exception as e:
            raise e

    def create_file(self, path):
        try:
            with open(path, 'w') as file:
                file.writelines('[')
        except Exception as e:
            raise e
        
    def get_datetime(self) -> datetime:
        return datetime.now()
    
    def datetime_parse_to_str(self, datetime) -> str:
        return datetime.strftime('%Y-%m-%d_%H-%M-%S')
    
    def check_if_solicitation_is_successfull(self, response) -> bool:
        try:
            if response.status_code == 200 and response.json():
                return True
            else:
                return False
        except Exception as e:
            raise e

    def write_satellite_in_file(self, satellite_list, path_to_file):
        try:
            with open(path_to_file, 'a') as file:
                json.dump({'time': self.datetime_parse_to_str(self.get_datetime()), 'satellites': satellite_list}, file, indent=4)
                file.writelines(',')
        except Exception as e:
            raise e
        
    def close_write_file(self, path_to_file):
        try:
            with open(path_to_file, 'r') as file:
                content = file.read()

            # Remover a vírgula extra no final do arquivo
            content = content.rstrip(',')
            content += ']'

            os.remove(path_to_file)

            # Salvar o conteúdo corrigido em um novo arquivo
            with open(path_to_file, 'w') as file:
                file.write(content)

        except Exception as e:
            raise e

    def save_logs_in_sheet(self, input_file):
        try:
            data = self.load_json(input_file)
            planilha, wb = self.create_sheet()
            data, planilha, satellites = self.create_headers(data, planilha)
            self.marking_sheet(data, planilha, satellites)
            output_file_name = 'dados'
            self.save_sheet(wb, output_file_name)
        except Exception as e:
            raise e

    def load_json(self, input_file):
        # Carrega o arquivo JSON
        with open(input_file, 'r') as file:
            data = json.load(file) # Carregando os dados do json em uma lista de dados
        return data
    
    def create_sheet(self):
        # Criando planilha
        wb = Workbook()
        planilha = wb.active 

        return planilha, wb
    
    def create_headers(self, data, planilha):
        satellites = []
        # Iterando sobre os dados para listar os tempos nos cabeçalhos
        for coluna, i in enumerate(data):
            planilha.cell(row=1, column=coluna+2, value=i['time'])
    
        # Iterando sobre os dados para procurar os satélites
        for linha, i in enumerate(data):
            # Iterando sobre os satélites para armazenar os respectivos campos 'satid' de maneira única
            for linha, j in enumerate(i['satellites']):
                if j['satid'] not in satellites:
                    satellites.append(j['satid'])
        
        # Escrevendo o id de cada satélite em uma coluna
        for linha, i in enumerate(satellites):
            planilha.cell(row=linha+2, column=1, value=i)

        return data, planilha, satellites

    def marking_sheet(self, data, planilha, satellites):
        # Definindo os valores para linha e coluna (a partir da posição [2][2])
        linha = coluna = 2

        # Iterando sobre os dados do json
        for i in data:
            # Definindo uma lista auxiliar que receberá os valores dos satélites localizados no tempo "X"
            lista = []

            # Iterando sobre os satélites do tempo X
            for j in i['satellites']:
                lista.append(j['satid']) # Adiciona o id a lista auxiliar
            
            # Iterando sobre todos os id's
            for j in satellites:
                if j in lista: # Se o id estiver na lista, significa que devemos marcar um "X" na célula
                    planilha.cell(row=linha, column=coluna, value="X")
                else:
                    planilha.cell(row=linha, column=coluna, value="--")
                linha+=1 # Andamos uma linha para baixo
            
            # Após sair do for, avançamos uma coluna e voltamos ao topo da tabela 
            # (linha 2 por cusa do cabeçalho)
            coluna+=1
            linha=2
        
    def save_sheet(self, wb, name='data'):
        wb.save(f'{name}.xlsx')